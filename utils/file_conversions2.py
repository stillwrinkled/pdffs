import io
import os
import zipfile
from pdf2docx import Converter as PDF2DOCXConverter
import fitz  # PyMuPDF for converting PDF to images (JPEG/PNG) and text extraction
from openpyxl import Workbook
from pptx import Presentation
from pptx.util import Inches
import camelot  # For extracting tables from PDF
from PyPDF2 import PdfReader
import tempfile
import tabula  # For extracting tables using Tabula
from pdfminer.high_level import extract_text  # For general text extraction using PDFMiner


def extract_with_camelot(file):
    try:
        # Try extracting tables using Camelot (stream mode first, lattice as fallback)
        print("Trying Camelot...")
        tables = camelot.read_pdf(file, pages='all', flavor='stream')

        if not tables or len(tables) == 0:
            print("No tables found in Camelot's stream mode, trying lattice mode...")
            tables = camelot.read_pdf(file, pages='all', flavor='lattice')

        if not tables or len(tables) == 0:
            raise ValueError("No tables found using Camelot.")
        return tables
    except Exception as e:
        print(f"Camelot failed: {e}")
        return None


def extract_with_tabula(file):
    try:
        print("Trying Tabula...")
        tables = tabula.read_pdf(file, pages='all', multiple_tables=True)
        if not tables or len(tables) == 0:
            raise ValueError("No tables found using Tabula.")
        return tables
    except Exception as e:
        print(f"Tabula failed: {e}")
        return None


def extract_with_pymupdf(file):
    try:
        print("Trying PyMuPDF...")
        doc = fitz.open(stream=file.read(), filetype="pdf")
        text = ""
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            text += page.get_text("text")
        if not text.strip():
            raise ValueError("No text found using PyMuPDF.")
        return text
    except Exception as e:
        print(f"PyMuPDF failed: {e}")
        return None


def extract_with_pdfminer(file):
    try:
        print("Trying PDFMiner...")
        text = extract_text(file)
        if not text.strip():
            raise ValueError("No text found using PDFMiner.")
        return text
    except Exception as e:
        print(f"PDFMiner failed: {e}")
        return None


def convert_pdf(file, format):
    output = io.BytesIO()

    # Reset file pointer to the start in case it has been read before
    file.seek(0)

    if format == 'docx':
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                temp_pdf.write(file.read())
                temp_pdf.close()
                output_file = io.BytesIO()
                converter = PDF2DOCXConverter(temp_pdf.name)
                converter.convert(output_file, start=0, end=None)
                converter.close()
                os.remove(temp_pdf.name)
                output_file.seek(0)
                output_filename = "converted_file.docx"
                return output_file, output_filename

        except Exception as e:
            print(f"Error during PDF to DOCX conversion: {e}")
            raise ValueError("Failed to convert PDF to DOCX.")

    elif format == 'xlsx':
        try:
            print("Attempting table extraction with Camelot")
            tables = extract_with_camelot(file)  # Try Camelot first
            if tables:
                wb = Workbook()
                ws = wb.active
                for table_num, table in enumerate(tables):
                    print(f"Processing table {table_num + 1} with Camelot")
                    for i, row in enumerate(table.df.itertuples(), 1):
                        for j, value in enumerate(row[1:], 1):
                            ws.cell(row=i, column=j, value=value)
                wb.save(output)
                output_filename = "converted_file.xlsx"
                output.seek(0)
                return output, output_filename

            print("Camelot failed, trying Tabula")
            tables = extract_with_tabula(file)  # Fallback to Tabula
            if tables:
                wb = Workbook()
                ws = wb.active
                for table_num, table in enumerate(tables):
                    print(f"Processing table {table_num + 1} with Tabula")
                    for i, row in enumerate(table.itertuples(), 1):
                        for j, value in enumerate(row[1:], 1):
                            ws.cell(row=i, column=j, value=value)
                wb.save(output)
                output_filename = "converted_file.xlsx"
                output.seek(0)
                return output, output_filename

            # If no tables found, fallback to text extraction (for manual table processing)
            print("Table extraction failed, attempting text extraction with PyMuPDF")
            text = extract_with_pymupdf(file) or extract_with_pdfminer(file)
            if text:
                print(f"Extracted text (first 200 characters): {text[:200]}")  # Print extracted text
                # Try processing the extracted text into rows/columns
                lines = text.strip().split('\n')
                wb = Workbook()
                ws = wb.active
                for i, line in enumerate(lines, 1):
                    columns = line.split()  # Split by spaces for simplicity
                    for j, value in enumerate(columns, 1):
                        ws.cell(row=i, column=j, value=value)
                wb.save(output)
                output_filename = "converted_file.xlsx"
                output.seek(0)
                return output, output_filename

            raise ValueError("Failed to extract tables or relevant text.")

        except Exception as e:
            print(f"Error during PDF to Excel conversion: {e}")
            raise ValueError("Failed to convert PDF to Excel. Please check if the PDF contains valid tables.")

    # Handling other formats (PPTX, JPEG, PNG) remains unchanged.

    output.seek(0)
    return output, None


def html_to_pdf(html_content):
    from xhtml2pdf import pisa
    output = io.BytesIO()
    pisa.CreatePDF(html_content, dest=output)
    output.seek(0)
    return output
